import requests
import json
import pandas as pd
from openpyxl import Workbook, load_workbook


'''
Obtains Manufacturer, Product, Notice of Compliance Date, and the Medical Ingridients
from the Health Canada API given the Drug Information Number (DIN).

Note that the NOC Database and the DGD database have different nomenclature
for the above terms, the code below follows the NOC database. 
'''

def get_drug_product_data(din):
    url = 'https://health-products.canada.ca/api/drug/drugproduct/?din=' + din + '&lang=en&type=json'
    response =  requests.get(url).json()
    #drug_code = response[0]['drug_code']
    return response[0] # only need to take the first in the list, as DIN is a primary key

def get_status(drug_code):
    url = 'https://health-products.canada.ca/api/drug/status/?id='+ str(drug_code) +'&lang=en&type=json'
    response = requests.get(url).json()
    return response['status']

def get_dosage_form(drug_code):
    url = 'https://health-products.canada.ca/api/drug/form/?id='+ str(drug_code) +'&lang=en&type=json'
    response_list = requests.get(url).json()
    forms = []
    for response in range(len(response_list)):
        forms.append(response_list[response]['pharmaceutical_form_name'])
    return forms

def get_schedule(drug_code):
    url = 'https://health-products.canada.ca/api/drug/schedule/?id='+ str(drug_code) +'&lang=en&type=json'
    response = requests.get(url).json()
    return response[0]['schedule_name']

## --------------------------------------------------------------------------------------------------------------------------------------------------    

wb = load_workbook('test_file1.xlsx')
ws = wb.active

din_list = []
for cell in ws['A']:
    if isinstance(cell.value, int):
        din = str(cell.value).zfill(8)
        din_list.append(din)

drug_df = pd.DataFrame(columns = ('DIN', 'Product Name', 'Company', 'Schedule', 'Status', 'Dosage Form(s)'))

for din in din_list:
    dp_data = get_drug_product_data(din)
    drug_code = str(dp_data['drug_code'])
    schedule = get_schedule(drug_code)
    dosage_form = get_dosage_form(drug_code)
    status = get_status(drug_code)

    new_row = {'DIN': din, 'Product Name': dp_data['brand_name'],
                'Company': dp_data['company_name'], 'Schedule': schedule,
                'Status': status, 'Dosage Form(s)': ', '.join(dosage_form)}

    drug_df = drug_df.append(new_row, ignore_index = True)

with pd.ExcelWriter('test_file1.xlsx', mode = 'a') as writer:
    drug_df.to_excel(writer, sheet_name = 'Final Sheet', index = False)

